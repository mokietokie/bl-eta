"""결과 DataFrame → CSV/xlsx 바이트 export.

Streamlit `st.download_button` data 인자에 직접 전달하는 용도.
헤더 굵게, 컬럼 폭은 데이터 최대 길이 기준 auto-fit (최소 8, 최대 40).
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def to_csv(df: pd.DataFrame) -> bytes:
    """UTF-8 BOM 포함 — Excel 한글 깨짐 방지."""
    return df.to_csv(index=False).encode("utf-8-sig")


def to_xlsx(df: pd.DataFrame, sheet_name: str = "bl-eta") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(df.columns)
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    for row in df.itertuples(index=False, name=None):
        ws.append([_cellify(v) for v in row])

    for col_idx, name in enumerate(headers, start=1):
        max_len = len(str(name))
        for v in df.iloc[:, col_idx - 1]:
            s = "" if v is None else str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max(max_len + 2, 8), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── shipments (마스터) Excel I/O ──────────────────────────────────────

# UI/엑셀 한글 헤더 ↔ DB 컬럼 매핑. 순서가 테이블/엑셀 컬럼 순서.
SHIPMENT_COLS: list[tuple[str, str]] = [
    ("smelter", "제련소"),
    ("origin", "출항지"),
    ("carrier", "선사"),
    ("vessel", "선명"),
    ("bl_no", "BL"),
    ("supply_tons", "공급물량(톤)"),
    ("initial_depart_date", "최초출항일"),
]
# 다운로드에만 포함되는 파생 컬럼 (업로드 시에는 무시)
SHIPMENT_DERIVED_KO: list[str] = ["국내 도착일", "전일 대비 변동"]


def shipments_to_xlsx(df: pd.DataFrame, sheet_name: str = "shipments") -> bytes:
    """마스터 df(한글 헤더 가정) 그대로 xlsx 바이트. to_xlsx와 동일 스타일."""
    return to_xlsx(df, sheet_name=sheet_name)


def shipments_from_xlsx(data: bytes) -> list[dict]:
    """엑셀 바이트 → shipments dict 리스트 (DB 컬럼명 키). 파생/알 수 없는 컬럼 무시.

    빈 BL 행 스킵. 공급물량 float 변환 실패 시 None.
    """
    raw = pd.read_excel(io.BytesIO(data), dtype=object)
    ko_to_db = {ko: db_ for db_, ko in SHIPMENT_COLS}
    out: list[dict] = []
    for _, row in raw.iterrows():
        rec: dict = {}
        for ko, db_ in ko_to_db.items():
            if ko not in raw.columns:
                continue
            v = row[ko]
            if v is None or (isinstance(v, float) and pd.isna(v)):
                rec[db_] = None
            else:
                rec[db_] = v
        bl = (rec.get("bl_no") or "")
        if isinstance(bl, float):
            bl = "" if pd.isna(bl) else str(int(bl)) if bl.is_integer() else str(bl)
        bl = str(bl).strip()
        if not bl:
            continue
        rec["bl_no"] = bl
        # 나머지 문자열 필드 trim
        for k in ("smelter", "origin", "carrier", "vessel", "initial_depart_date"):
            if rec.get(k) is not None:
                rec[k] = str(rec[k]).strip() or None
        out.append(rec)
    return out


def _cellify(v):
    """None/NaN → 빈 문자열. 나머지는 그대로."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return v
